import pandas as pd
import os
import openpyxl
def cut(address_name,target_path,percent_text,percent_verify):
    # address_name：地址和文件名 target_path：指定路径 percent_text:测试集占总数据集比例 percent_verify：验证集占训练集比例
    # cut.cut('D:\\分离器\\cut\\data.xlsx', 'D:\\Excel文件', 0.2, 0.1)
    data=pd.read_excel(address_name,engine='openpyxl')
    text_data=data.sample(frac=percent_text)
    train_data=data.drop(text_data.index)
    verify_data=train_data.sample(frac=percent_verify)
    print('data=\n',data)
    print('text_data=\n',text_data)
    print('train_data=\n',train_data)


    path=target_path
    os.chdir(path)
    list1=[]
    for index,col in text_data.iterrows():
        list1.append(col)
    print(list1)
    df=pd.DataFrame(list1)
    df.to_excel('text_data.xlsx',index=True)

    list2=[]
    for index,col in train_data.iterrows():
        list2.append(col)
    print(list2)
    df=pd.DataFrame(list2)
    df.to_excel('train_data.xlsx',index=True,)

    list3=[]
    for index,col in verify_data.iterrows():
        list3.append(col)
    print(list1)
    df=pd.DataFrame(list3)
    df.to_excel('verify_data.xlsx',index=True)


    path = target_path
    os.chdir(path)
    workbook = openpyxl.load_workbook('text_data.xlsx')
    sheet = workbook.active
    sheet.delete_cols(idx=1)
    workbook.save('{}\\text_data.xlsx'.format(target_path))#生成测试集文件


    workbook = openpyxl.load_workbook('train_data.xlsx')
    sheet = workbook.active
    sheet.delete_cols(idx=1)
    workbook.save('{}\\train_data.xlsx'.format(target_path))#生成训练集文件


    workbook = openpyxl.load_workbook('verify_data.xlsx')
    sheet = workbook.active
    sheet.delete_cols(idx=1)
    workbook.save('{}\\verify_data.xlsx'.format(target_path))#生成验证集文件
